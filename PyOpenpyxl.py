#-------------------------------------------------------------------------------
# Name:        PyOpenpyxl.py
# Purpose:
#
# Author:      Shyed Shahriar Housaini
#
# Created:     24/10/2019
# Copyright:   (c) Shyed Shahriar Housaini 2019
# Licence:     <your licence: Shyed Shahriar Housaini's terms and conditions.>
#-------------------------------------------------------------------------------

print(""" --------- Start of written code ----------
import openpyxl
import os
  ======= End of written code ========== End of written code ======= Output ======Output ====== """)

import openpyxl
import os

print(""" --------- Start of written code ----------
## os.chdir('C:\\Users\\HP\\Desktop') ## For office computer. ##

os.chdir('C:\\Users\\HP\\Desktop') ## For home computer. ##
## Always use \\ for avoiding any error in windows.
## use / or // for linux
  ======= End of written code ========== End of written code ======= Output ======Output ====== """)

## os.chdir('C:\\Users\\HP\\Desktop') ## For office computer. ##

os.chdir('C:\\Users\\HP\\Desktop') ## For home computer. ##
## Always use \\ for avoiding any error in windows.
## use / or // for linux

print(""" --------- Start of written code ----------
dirpath = os.getcwd()
print("current directory is : " + dirpath)
  ======= End of written code ========== End of written code ======= Output ======Output ====== """)

dirpath = os.getcwd()
print("current directory is : " + dirpath)

print(""" --------- Start of written code ----------

foldername = os.path.basename(dirpath)
print("Directory name is : " + foldername)

scriptpath = os.path.realpath(__file__)
print("Script path is : " + scriptpath)
print ("Current openpyxl version :" , openpyxl.__version__)

  ======= End of written code ========== End of written code ======= Output ======Output ====== """)

foldername = os.path.basename(dirpath)
print("Directory name is : " + foldername)

scriptpath = os.path.realpath(__file__)
print("Script path is : " + scriptpath)

print ("Current openpyxl version :" , openpyxl.__version__)

print(""" --------- Start of written code ----------
wb=openpyxl.load_workbook('C:\\Users\\HP\\Desktop\\Testopenpyxl.xlsx')
# type(wb) or print (wb) #
print (wb)
SN=wb.get_sheet_names()
print('sheet_names : '  , SN)
  ======= End of written code ========== End of written code ======= Output ======Output ====== """)

wb=openpyxl.load_workbook('C:\\Users\\HP\\Desktop\\Testopenpyxl.xlsx')
# type(wb) or print (wb) #
print (wb)

SN=wb.get_sheet_names()
print('sheet_names : '  , SN)

print(""" --------- Start of written code ----------
sheet= wb.get_sheet_by_name('Sheet1')
# type() or print () #
print (sheet)
  ======= End of written code ========== End of written code ======= Output ======Output ====== """)

sheet= wb.get_sheet_by_name('Sheet1')

# type() or print () #
print (sheet)

print(""" --------- Start of written code ----------
sheet['A1'].value
sheeta1=sheet['A1'].value
print(sheeta1)
sheet['A4'].value
sheeta4=sheet['A4'].value
print(sheeta4)
sheet['A7'].value
sheeta7=sheet['A7'].value
print(sheeta7)

  ======= End of written code ========== End of written code ======= Output ======Output ====== """)

sheet['A1'].value
sheeta1=sheet['A1'].value
print(sheeta1)
sheet['A4'].value
sheeta4=sheet['A4'].value
print(sheeta4)
sheet['A7'].value
sheeta7=sheet['A7'].value
print(sheeta7)

print(""" --------- Start of written code ----------
sheet['B1'].value
sheetB1=sheet['B1'].value
print(sheetB1)
sheet['B4'].value
sheetB4=sheet['B4'].value
print(sheetB4)
sheet['B7'].value
sheetB7=sheet['B7'].value
print(sheetB7)

  ======= End of written code ========== End of written code ======= Output ======Output ====== """)
sheet['B1'].value
sheetB1=sheet['B1'].value
print(sheetB1)
sheet['B4'].value
sheetB4=sheet['B4'].value
print(sheetB4)
sheet['B7'].value
sheetB7=sheet['B7'].value
print(sheetB7)

print(""" ~!@#$%^&* --------- Start of written code ---------- ~!@#$%^&*
sheet['C1'].value
sheetC1=sheet['C1'].value
print(sheetC1)
sheet['C4'].value
sheetC4=sheet['C4'].value
print(sheetC4)
sheet['C7'].value
sheetC7=sheet['C7'].value
print(sheetC7)

~!@#$%^&*   ======= End of written code ========== End of written code ======= Output ======Output ======  ~!@#$%^&* """)
sheet['C1'].value
sheetC1=sheet['C1'].value
print(sheetC1)
sheet['C4'].value
sheetC4=sheet['C4'].value
print(sheetC4)
sheet['C7'].value
sheetC7=sheet['C7'].value
print(sheetC7)


print(""" --------- Start of written code ----------
sheet['C11'].value = 43
wb.save('C:\\Users\\HP\\Desktop\\2ndTestopenpyxl.xlsx')
## can not save into currently open files.

  ======= End of written code ========== End of written code ======= Output ======Output ====== """)
sheet['C11'].value = 43
wb.save('C:\\Users\\HP\\Desktop\\2ndTestopenpyxl.xlsx')
## can not save into currently open files.


print(""" --------- Start of written code ----------
STL=sheet.title
print(STL) ## current sheet name

  ======= End of written code ========== End of written code ======= Output ======Output ====== """)
STL=sheet.title
print(STL) ## current sheet name

print(""" --------- Start of written code ----------
sheet.title= "My wb sheet"
wb.save('C:\\Users\\HP\\Desktop\\2ndTestopenpyxl.xlsx')

  ======= End of written code ========== End of written code ======= Output ======Output ====== """)

sheet.title= "My wb sheet"
wb.save('C:\\Users\\HP\\Desktop\\2ndTestopenpyxl.xlsx')

print(""" --------- Start of written code ----------
STL2=sheet.title
print(STL2) ## current sheet name

  ======= End of written code ========== End of written code ======= Output ======Output ====== """)
STL2=sheet.title
print(STL2) ## current sheet name

print(""" --------- Start of written code ----------
row_column= sheet.cell(row=1, column=1).value
print(row_column)
  ======= End of written code ========== End of written code ======= Output ======Output ====== """)
row_column= sheet.cell(row=1, column=1).value
print(row_column)

print(""" --------- Start of written code ----------
row_column113= sheet.cell(row=11, column=3).value

## (row=11, column=3) = C11
print(row_column113)

  ======= End of written code ========== End of written code ======= Output ======Output ====== """)

row_column113= sheet.cell(row=11, column=3).value

## (row=11, column=3) = C11
print(row_column113)

print(""" --------- Start of written code ----------
for i in range (1,5):
    print(sheet.cell (row=i, column=3).value)
  ======= End of written code ========== End of written code ======= Output ======Output ====== """)

for i in range (1,5):
    print(sheet.cell (row=i, column=3).value)

print(""" --------- Start of written code ----------
## for knowing maximum numbers of rows and columns
##sheet.max_row
##sheet.max_column
  ======= End of written code ========== End of written code ======= Output ======Output ====== """)
## for knowing maximum numbers of rows and columns
##sheet.max_row
##sheet.max_column


print(""" --------- Start of written code ----------
print(sheet.max_row)
print(sheet.max_column)

print(openpyxl.utils.get_column_letter(1))

  ======= End of written code ========== End of written code ======= Output ======Output ====== """)

print(sheet.max_row)
print(sheet.max_column)

print(openpyxl.utils.get_column_letter(1))



print(openpyxl.utils.get_column_letter(1))


print(""" --------- Start of written code ----------
## print(openpyxl.cell.get_column_letter(1))
## does not work rather use
## print(openpyxl.utils.get_column_letter(1)) ##

  ======= End of written code ========== End of written code ======= Output ======Output ====== """)
## print(openpyxl.cell.get_column_letter(1))
## does not work rather use
## print(openpyxl.utils.get_column_letter(1)) ##

print(""" --------- Start of written code ----------
print(openpyxl.utils.get_column_letter(26))
print(openpyxl.utils.get_column_letter(27))

  ======= End of written code ========== End of written code ======= Output ======Output ====== """)
print(openpyxl.utils.get_column_letter(26))
print(openpyxl.utils.get_column_letter(27))

print(""" --------- Start of written code ----------
print(openpyxl.utils.column_index_from_string('AA'))
print(openpyxl.utils.column_index_from_string('AAA'))
  ======= End of written code ========== End of written code ======= Output ======Output ====== """)

print(openpyxl.utils.column_index_from_string('AA'))
print(openpyxl.utils.column_index_from_string('AAA'))




print("""

--------- Start of written code ---------- --------- Start of written code ----------

wb.create_sheet(title="My 2nd sheet", index=1)
wb.save('C:\\Users\\HP\\Desktop\\2ndTestopenpyxl.xlsx')

 ======= End of written code ========== End of written code ======= Output ======""")

wb.create_sheet(title="My 2nd sheet", index=1)
wb.save('C:\\Users\\HP\\Desktop\\2ndTestopenpyxl.xlsx')

print("""

--------- Start of written code ---------- --------- Start of written code ----------

sheet.row_dimensions[1].height =71
sheet.row_dimensions[11].height =77
sheet.column_dimensions['E'].width = 43
sheet.column_dimensions['D'].width = 34
wb.save('C:\\Users\\HP\\Desktop\\2ndTestopenpyxl.xlsx')

======= End of written code ========== End of written code =======
output ======
  """)
sheet.row_dimensions[1].height =71
sheet.row_dimensions[11].height =77
sheet.column_dimensions['E'].width = 43
sheet.column_dimensions['D'].width = 34
wb.save('C:\\Users\\HP\\Desktop\\2ndTestopenpyxl.xlsx')

print("""

--------- Start of written code ---------- --------- Start of written code ----------
print(sheet.row_dimensions[1].height)
print(sheet.row_dimensions[2].height)
print(sheet.column_dimensions['A'].width)
print(sheet.column_dimensions['B'].width)

======= End of written code ========== End of written code =======
output ======
  """)

print(sheet.row_dimensions[1].height)
print(sheet.row_dimensions[2].height)
print(sheet.column_dimensions['A'].width)
print(sheet.column_dimensions['B'].width)

print("""

--------- Start of written code ---------- --------- Start of written code ----------
reload the workbook
wb=openpyxl.load_workbook('C:\\Users\\HP\\Desktop\\2ndTestopenpyxl.xlsx')
sheet=wb.get_sheet_by_name('My wb Sheet')
type(sheet)
print(sheet)

======= End of written code ========== End of written code =======

output ======
  """)

##reload the workbook
wb=openpyxl.load_workbook('C:\\Users\\HP\\Desktop\\2ndTestopenpyxl.xlsx')
sheet=wb.get_sheet_by_name('My wb sheet')
## type(sheet) ## this does not work
print(sheet)

print("""

--------- Start of written code ---------- --------- Start of written code ----------
from openpyxl.styles import Font
sheet['B1'].font=Font(sz=14, bold=True, italic=True)
wb.save('C:\\Users\\HP\\Desktop\\2ndTestopenpyxl.xlsx')

======= End of written code ========== End of written code =======

output ======
  """)
from openpyxl.styles import Font
sheet['B1'].font=Font(sz=14, bold=True, italic=True)
sheet['B1'].font=Font(sz=14, bold=True, italic=True)
wb.save('C:\\Users\\HP\\Desktop\\2ndTestopenpyxl.xlsx')

print("""

--------- Start of written code ---------- --------- Start of written code ----------
import random
## Create sheet by
wb.create_sheet(title="My 3rd sheet", index=2)
## OR Create sheet by
sheet=wb.create_sheet('My 4th sheet')

for i in range (1, 111, 2):
    sheet['A' + str (i)].value = random.randint(1,1111)

wb.save('C:\\Users\\HP\\Desktop\\2ndTestopenpyxl.xlsx')

======= End of written code ========== End of written code =======

output ======
  """)

import random
## Create sheet by
wb.create_sheet(title="My 3rd sheet", index=2)
## OR Create sheet by
sheet=wb.create_sheet('My 4th sheet')

for i in range (1, 111, 2):
    sheet['A' + str (i)].value = random.randint(1,1111)

wb.save('C:\\Users\\HP\\Desktop\\2ndTestopenpyxl.xlsx')

for i in range (1, 111):
    sheet['B' + str (i)].value = random.randint(1,9999)

wb.save('C:\\Users\\HP\\Desktop\\2ndTestopenpyxl.xlsx')

print("""

--------- Start of written code ---------- --------- Start of written code ----------
refObj= openpyxl.chart.Reference(sheet(1,1),(10,1))
wb.save('C:\\Users\\HP\\Desktop\\2ndTestopenpyxl.xlsx')
print(refObj)


======= End of written code ========== End of written code =======

output ======
  """)

import os

from selenium import webdriver

borwser = webdriver.Chrome()

browser.get('https://www.youtube.com')

elem = browser.find_element_by_css_selector('#description > yt-formatted-string > a:nth-child(1)')

elem.texy

elem.click()

browser.quit()

##refObj= openpyxl.chart.Reference(sheet(1,1),(10,1))
##wb.save('C:\\Users\\HP\\Desktop\\2ndTestopenpyxl.xlsx')
##print(refObj)



print("""

--------- Start of written code ---------- --------- Start of written code ----------


======= End of written code ========== End of written code =======

output ======
  """)



print("""

--------- Start of written code ---------- --------- Start of written code ----------


======= End of written code ========== End of written code =======

output ======
  """)




